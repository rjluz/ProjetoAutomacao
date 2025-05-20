#Importa o driver do navegador (Chrome, Firefox etc.) para automação
from selenium import webdriver
# Permite simular pressionar teclas do teclado (como Enter, Tab etc.)
from selenium.webdriver.common.keys import Keys
# Permite localizar elementos na página (por ID, classe, nome, XPath etc.)
from selenium.webdriver.common.by import By
#a linha abaixo importa biblioteca para trabalhar com a planilha
from openpyxl import load_workbook
#importa a hora em que a aplicação foi executada
from datetime import datetime
#abre a planilha automaticamente depois da execução do programa
import os
#biblioteca pra criação da tela gráfica
from tkinter import *

class Aplicacao:
    def __init__(self):
        #cria o layout da aplicação
        self.layout = Tk()
        #cria o titulo da janela
        self.layout.title('Captador de Temperatura de São Gonçalo')
        #cria o tamanho da tela
        self.layout.geometry('480x120')
        #cria um pequeno quadrado (tela)
        self.tela = Frame(self.layout)
        #cria a descrição do botao
        self.descricao = Label(self.tela, text='atualizar previsão na planilha')
        #cria o botão   
        self.exportar = Button(self.tela, text='Gerar Arquivo',command=exportar_arquivo)
        # Posiciona o quadro na janela
        self.tela.pack()
        # Posiciona o texto (label) dentro do quadro
        self.descricao.pack()
        # Posiciona o botão dentro do quadro
        self.exportar.pack()
        #Mantém a janela da aplicação aberta e ouve eventos (cliques, teclas, etc).
        self.layout.mainloop()

def exportar_arquivo():
    #cria a variavel que pega a hora em que rodou aplicação
    hora_execucao = datetime.now().strftime('%H:%M')
    print(hora_execucao)

    #abre o navegador
    navegador = webdriver.Chrome()
    #Abre no link q eu inseri
    navegador.get('https://www.climatempo.com.br/previsao-do-tempo/cidade/325/saogoncalo-rj')

    # 1. Captura o texto usando XPath 
    elemento1 = navegador.find_element(By.XPATH, '//*[@id="mainContent"]/div[7]/div[3]/div[1]/div[2]/div[1]/div/h1')
    #transforma o arquivo x.path em texto
    data = elemento1.text

    # Extrai só a data (sem usar biblioteca extra)
    for palavra_data in data.split():     
     if "/" in palavra_data:
        data = palavra_data
        break

    print(palavra_data)
    
    elemento2 = navegador.find_element(By.XPATH,'//*[@id="mainContent"]/div[7]/div[3]/div[1]/div[2]/div[2]/div[3]/div[1]/ul/li[1]/div/p')
    temperatura = elemento2.text

   # Divide a string de temperatura em duas partes: mínimo e máximo, e remove o símbolo de grau (°)
    partes = temperatura.split()
    min_temp = partes[0].replace("°", "")
    max_temp = partes[1].replace("°", "")

    print(f'Min: {min_temp}C')
    print(f'Max {max_temp}C')

    elemento3 = navegador.find_element(By.XPATH,'//*[@id="mainContent"]/div[7]/div[3]/div[1]/div[2]/div[2]/div[3]/div[1]/ul/li[4]/div/p')
    umidade = elemento3.text

    partes = umidade.split()
    min_umid = partes[0]
    max_umid = partes[1]

    print(f'Min {min_umid}')
    print(f'Max {max_umid}')


    #a linha abaixo abre o arquivo
    arquivo = load_workbook('Planilhatemp.xlsx')
    #abre o arquivo na folha 'TEMP'
    planilha = arquivo['TEMP']

    linha = planilha.max_row + 1  # calcula a próxima linha vazia

    planilha.cell(row=linha, column=1).value = palavra_data
    planilha.cell(row=linha, column=2).value = hora_execucao  
    planilha.cell(row=linha, column=3).value = f"Min:{min_temp}C - Max:{max_temp}C"
    planilha.cell(row=linha, column=4).value = f'Min:{min_umid} - Max:{max_umid}'

    #salva o arquivo
    arquivo.save('Planilhatemp.xlsx')

    #abre o excell com todas as informações salvas
    os.startfile('Planilhatemp.xlsx')   
    
    #fecha o navegador
    navegador.quit()

#É a variavel da tela da aplicação onde tudo da linha acima vai acontecer
tl = Aplicacao()

print("Planilha atualizada com sucesso")